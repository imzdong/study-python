import requests
import openpyxl
import config
import time
import random
from bs4 import BeautifulSoup
import re
import os
import urllib
from jinja2 import Environment, FileSystemLoader
import json

'''
关键字解释：
cooike帮我们绕过登录过程
fakeid是目标公众号的唯一标识符
user-agent可以模拟浏览器请求至此信息获取部分完成，下面开始开始代码部分。
'''

headers = {
    "cookie": "appmsglist_action_3094473706=card; ua_id=rTisf6no8nQ6Z2EpAAAAAE0nkbQrFlZq1A7-67LqGnU=; wxuin=22044998669898; uuid=e97fca24606f7c01a7fb3cd1ce3c8ae2; _clck=6s741v|1|fnt|0; rand_info=CAESIIFwJvXVCYaaNxOrlF5oVVrHO7PD4l0NSbFBr60xeihz; slave_bizuin=3094473706; data_bizuin=3094473706; bizuin=3094473706; data_ticket=PcUMElrHiY0jd+fBQkn8WWDrd7fNtben8VE3VYLs1YLSrcipat/O2soiIuY1LUeh; slave_sid=UFlGeE85Qm5nMzhsY1dQcXhZUV83ZWJfRHM0Z3NsSExYUkg4eE5ndjVqOXpDSVU4TE9sMXNHbHRRTXRRc3dBWDAyeHVyd2ZlTXBzbnI3V1BLSEV4RmFsNGlfSFpxdVo4RTJ4VnZIMVdxVG1iVzlNd2Y4bVVrcW9uQ2pYNWdPaWZuN0hMNDdTSWdEdHBIZUNq; slave_user=gh_195ed1058a3c; xid=8e39e241dcd2b96bb3d869f1049417a6; mm_lang=zh_CN; _clsk=hxabmm|1722044983139|3|1|mp.weixin.qq.com/weheat-agent/payload/record",
    "X-Requested-With": "XMLHttpRequest",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36"
}

url = 'https://mp.weixin.qq.com/cgi-bin/appmsg'
fad = 'MzUzMjY0NDY4Ng%3D%3D'
token = '223369894'
count = 5
# 爬不同公众号只需要更改 fakeid 1857985110
rootPath = 'D:\\WorkSpace\\idea\\python\\wechat\\'
wxlistfile = rootPath + 'wxlist.xlsx'
htmlPath = rootPath + 'html'



def getwxlist(wxid):

    begin = "0"
    params = {
        "action": "list_ex",
        "begin": begin,
        "count": count,
        "fakeid": wxid,
        "type": "9",
        "token": token,
        "lang": "zh_CN",
        "f": "json",
        "ajax": "1"
    }
    # 会封锁的，但是隔一段时间就会解封了。拉取间隔长一点，像同一个公众号的不同文章间隔30s获取一次，不同公众号间隔1分钟，这样就基本不会被封。虽然总体时间长了点，但是胜在稳定
    # 获取历史抓取记录
    wb = openpyxl.load_workbook(wxlistfile)
    ws = wb.active
    # 在不知道公众号有多少文章的情况下，使用while语句
    # 也方便重新运行时设置页数
    i = 0
    # 新增抓取计数
    j = 0
    flag = False
    while True:
        begin = i * count
        params["begin"] = str(begin)
        resp = requests.get(url, headers=headers, params=params)
        # 微信流量控制, 退出
        if resp.json()['base_resp']['ret'] == 200013:
            print("frequencey control, stop at {}".format(str(begin)))
            wb.save(wxlistfile)
            time.sleep(random.randint(60*60, 62*60))
            continue
        # 如果返回的内容中为空则结束
        if len(resp.json()['app_msg_list']) == 0:
            print("all ariticle parsed")
            break
        msg = resp.json()
        if "app_msg_list" in msg:
            for item in msg["app_msg_list"]:
                op = json.dumps(item)
                # printing result as string
                print("final string = ", op)
                row = [item['title'], item['link'], op]
                ws.append(row)
                j += 1
            print(f"第{i}页爬取成功\n")
        if flag == True:
            print(f"新增页面抓取已完成,{j}篇文章已添加\n")
            break
        # 翻页
        i += 1
        # 随机暂停几秒，避免过快的请求导致过快的被查到
        time.sleep(random.randint(1*60, 3*60))
    wb.save(wxlistfile)

# 解析url
def get_beautiful_soup(url1):
    res = requests.get(url1)
    res.encoding = 'utf-8'
    html = res.text
    return BeautifulSoup(html, "html.parser")

def get_content(curl):
    data = {}
    soup = get_beautiful_soup(curl)
    if soup.find('h1', attrs={'id': 'activity-name'}) == None:
        data['title'] = "Nonetipycal Type"
    else:
        # Get Document Title
        data['title'] = soup.find('h1', attrs={'id': 'activity-name'}).string.strip()
        # Get the publish date parameter
        dateframe = re.findall(r'var ct\s*=\s*.*\d{10}', str(soup))
        # split the parameter as a list
        date = re.split('"', dateframe[0])
        #format the publish date
        data['date'] = time.strftime("%Y-%m-%d",time.localtime(int(date[1])))
        data['time'] = time.strftime("%Y-%m-%d %H:%M",time.localtime(int(date[1])))
        # Get the content
        data['content'] = soup.find('div', attrs={'id': 'js_content'})
        #this makes a list of bs4 element tags img
        data['images'] = [img for img in soup.find('div', attrs={'id': 'js_content'}).find_all('img')]
    return data

imgRootPath = 'https://mp.weixin.qq.com'

def saveData(curl):
    data = get_content(curl)
    image_links = []
    if data.get('images') != None:
        if len(data['images']) != 0:
            # compile our unicode list of image links
            image_links = [each.get('data-src') for each in data['images']]
            # create img folder
            # imgFolder = validateTitle (data['time'].split(' ')[0])
            imgFolder = validateTitle(data['date'])
            imgDst = os.path.join(htmlPath, 'imgs', imgFolder).replace('\\', '/')
            if not os.path.exists(imgDst):
                os.makedirs(imgDst)  # make directory
            for each in image_links:
                imagename = each.split('/')[-2]
                # convert abs address
                if not each.startswith("http"):
                    each = imgRootPath + each
                # save images
                download_image(each, os.path.join(imgDst, imagename).replace('\\', '/'))
                # join a file name with title and data & replace ilegal tags.
        filename = validateTitle(data['title'] + data['date'] + '.html')
        # replace ilegal tags
        saveDst = os.path.join(htmlPath, filename).replace('\\', '/')

        cleanSoup = data['content']
        if len(image_links) != 0:
            for each in image_links:
                imagename = each.split('/')[-2]
                srcNew = "./imgs/" + imgFolder + "/" + imagename

                cleanSoup.find('img', {'data-src': each})['src'] = srcNew
                # cleanSoup = BeautifulSoup(str(originalSoup).replace(old, new))
                # cleanSoup = BeautifulSoup(str(cleanSoup).replace(each, srcNew),'html.parser')
        cleanSoup['style'] = 'visibility: initial'
        # Format the parsed html file
        htmlcontent = _render_file(
                'article.html',
                {'title': filename, 'content': cleanSoup.prettify()}
            )
        # print(htmlcontent)

        with open(saveDst, "wt", encoding=("utf-8")) as f:
            f.write(htmlcontent)

        savetolist(curl, data['title'], saveDst, data['date'])
    else:
        saveDst = "None"
        saveDate = time.strftime('%Y-%m-%d')
        savetolist(curl, data['title'], saveDst, saveDate)


def download_image(imageUrl, local_filename):
    # 发送GET请求
    print(f"Image downloaded to {imageUrl}")
    response = requests.get(imageUrl, stream=True)

    # 检查响应状态码是否为200（成功）
    if response.status_code == 200:
        # 打开文件，使用'wb'模式来写入二进制数据
        with open(local_filename, 'wb') as file:
            # 写入图片数据
            for chunk in response.iter_content(chunk_size=8192):
                file.write(chunk)

        print(f"Image downloaded to {local_filename}")
    else:
        print("Failed to download the image.")

'''
html_new_content = _render_file(
                'article.html',
                {'title': filename, 'content': html_content}
            )
'''


def _render_file(template_name: str, context: dict) -> str:
    """
    生成 html 文件
    """
    # 设置模板文件夹
    file_loader = FileSystemLoader(rootPath+'templates')
    env = Environment(loader=file_loader)

    # 加载模板
    template = env.get_template(template_name)

    # 渲染模板
    rendered_html = template.render(context)
    return rendered_html

def validateTitle(title):
    rstr = r"[\/\\\:\*\?\"\<\>\|]"  # '/ \ : * ? " < > |'
    new_title = re.sub(rstr, "_", title)  # replace to _
    return new_title


def savetolist(curl, ctitle, lcfile, date):
    lclist = rootPath + '/lclist.xlsx'
    wb = openpyxl.load_workbook(lclist)
    ws = wb.active
    row = [curl, ctitle, lcfile, date]
    ws.append(row)
    wb.save(lclist)

def doGet():
    wb = openpyxl.load_workbook(wxlistfile, True)
    # 获取活动工作表
    ws = wb.active
    # 遍历每一行
    for row in ws.iter_rows(values_only=True):
        print(row[1])
        saveData(row[1])
        time.sleep(random.randint(1, 5))

def savetolistTest():
    lclist = rootPath + '/lclist-1.xlsx'
    wb = openpyxl.load_workbook(lclist)
    ws = wb.active
    row = [92, 92, 92, 92]
    ws.append(row)
    wb.save(lclist)
    row = [93, 93, 93, 93]
    ws.append(row)
    wb.save(lclist)

if __name__ == '__main__':
    #ar_url = 'http://mp.weixin.qq.com/s?__biz=MzUzMjY0NDY4Ng==&mid=2247502370&idx=1&sn=9be76a798bf0ca93af3d6d7c01115b24&chksm=fab29c03cdc5151553f6ce99e4b78a17761e589fba2b07824e73a4c29f2318e2d621c86d0773#rd'
    #saveData(ar_url)
    #doGet()
    #savetolistTest()
    getwxlist(fad)

